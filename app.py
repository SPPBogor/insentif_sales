from flask import (
    Flask, render_template, render_template_string, request,
    redirect, url_for, flash, send_file, session, abort
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, asc
from datetime import datetime
import pandas as pd
import io, os, re
from functools import lru_cache


# ===================== APP CONFIG =====================
app = Flask(__name__)
app.config["SECRET_KEY"] = "change-this-in-prod"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///app.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ===================== MODELS =====================
def latest_bulan(Model):
    try:
        m = db.session.query(func.max(Model.bulan)).scalar()
        return (m or "").strip()
    except Exception:
        return ""
    
def get_threshold(kind, divisi):
    row = Threshold.query.filter_by(kind=kind, divisi=divisi).first() \
       or Threshold.query.filter_by(kind=kind, divisi=None).first()
    return float(row.percent) if row else 100.0  # simpan 85 untuk 85%

def is_lolos(kind, divisi, target, actual):
    try:
        t = float(target or 0); a = float(actual or 0)
    except:
        t, a = 0.0, 0.0
    if t <= 0:
        return False
    return (a / t) * 100.0 >= get_threshold(kind, divisi)

# <- ini yang penting
app.jinja_env.globals.update(is_lolos=is_lolos)

class Threshold(db.Model):
    __tablename__ = "thresholds"
    id = db.Column(db.Integer, primary_key=True)
    kind = db.Column(db.String(50))      # "password_awal" | "penjualan" | "tagihan"
    divisi = db.Column(db.String(20))    # "MP" | "AVIAN"
    percent = db.Column(db.Float, nullable=True)
    updated_at = db.Column(db.DateTime, default=func.now(), onupdate=func.now())

class PasswordAwal(db.Model):
    __tablename__ = "password_awal"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    bulan = db.Column(db.String(7), nullable=False)  # YYYY-MM
    target_amount = db.Column(db.Float, nullable=False, default=0)
    actual_amount = db.Column(db.Float, nullable=False, default=0)

    __table_args__ = (
        db.UniqueConstraint("salesman_id", "bulan", name="uq_password_awal"),)

class Salesman(db.Model):
    __tablename__ = "salesman"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    area = db.Column(db.String(80), nullable=True)
    divisi = db.Column(db.String(10), nullable=False)   # MP | AVIAN
    kategori = db.Column(db.String(12), nullable=False) # Bronze | Silver | Gold | Platinum

class TargetPenjualan(db.Model):
    __tablename__ = "target_penjualan"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    bulan = db.Column(db.String(7), nullable=False)  # YYYY-MM
    target_amount = db.Column(db.Float, nullable=False, default=0)
    actual_amount = db.Column(db.Float, nullable=False, default=0)

class TargetTagihan(db.Model):
    __tablename__ = "target_tagihan"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    bulan = db.Column(db.String(7), nullable=False)  # YYYY-MM
    target_amount = db.Column(db.Float, nullable=False, default=0)
    actual_amount = db.Column(db.Float, nullable=False, default=0)

class NOO(db.Model):
    __tablename__ = "noo"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    bulan = db.Column(db.String(7), nullable=False)  # YYYY-MM
    target_ca = db.Column(db.Integer, nullable=True, default=0)
    actual_ca = db.Column(db.Integer, nullable=True, default=0)
    jumlah_toko_noo = db.Column(db.Integer, nullable=True, default=0)  # NOO/LOST count
    target_fl = db.Column(db.Integer, nullable=True, default=0)   # MP only
    actual_fl = db.Column(db.Integer, nullable=True, default=0)   # MP only
    target_ao = db.Column(db.Integer, nullable=True, default=0)   # AVIAN only
    actual_ao = db.Column(db.Integer, nullable=True, default=0)   # AVIAN only

class Call(db.Model):
    __tablename__ = "call"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    bulan = db.Column(db.String(7), nullable=False)  # YYYY-MM
    target_call = db.Column(db.Integer, nullable=True, default=0)
    actual_call = db.Column(db.Integer, nullable=True, default=0)

class InsentifRate(db.Model):
    """
    Rate insentif umum (editable via UI).
    - jenis: 'penjualan' | 'tagihan' | 'ca' | 'noo_item' | 'fl' | 'ao' | 'call'
    - divisi/kategori dipakai untuk penjualan & tagihan; kosong untuk jenis lain.
    """
    __tablename__ = "insentif_rate"
    id = db.Column(db.Integer, primary_key=True)
    jenis = db.Column(db.String(20), nullable=False)
    divisi = db.Column(db.String(10), nullable=True)       # MP | AVIAN | None
    kategori = db.Column(db.String(12), nullable=True)     # Bronze | Silver | Gold | Platinum | None
    amount = db.Column(db.Integer, nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)
    __table_args__ = (
        db.UniqueConstraint("jenis", "divisi", "kategori", name="uq_insentif_key"),
    )

class InsentifOverride(db.Model):
    """
    Rate insentif khusus per-salesman, opsional per-bulan.
    - jenis: seperti di InsentifRate
    - bulan: 'YYYY-MM' atau NULL (berlaku untuk semua bulan)
    """
    __tablename__ = "insentif_override"
    id = db.Column(db.Integer, primary_key=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey("salesman.id"), nullable=False)
    jenis = db.Column(db.String(20), nullable=False)
    bulan = db.Column(db.String(7), nullable=True)   # YYYY-MM
    amount = db.Column(db.Integer, nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)
    __table_args__ = (
        db.UniqueConstraint("salesman_id", "jenis", "bulan", name="uq_override_key"),
    )

class User(db.Model):
    __tablename__ = "users"
    id          = db.Column(db.Integer, primary_key=True)
    username    = db.Column(db.String(64), unique=True, nullable=False, index=True)
    full_name   = db.Column(db.String(128), nullable=False)
    role        = db.Column(db.String(16), nullable=False, default="admin")  # admin|superadmin
    password_hash = db.Column(db.String(255), nullable=False)
    is_active   = db.Column(db.Boolean, default=True, nullable=False)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def set_password(self, raw):
        from werkzeug.security import generate_password_hash
        self.password_hash = generate_password_hash(raw)

    def check_password(self, raw):
        from werkzeug.security import check_password_hash
        return check_password_hash(self.password_hash, raw)

# --- pastikan ini GLOBAL (tidak ke-indent di dalam class) ---
def ensure_super_admin():
    if not User.query.filter_by(role="superadmin").first():
        sa = User(username="superadmin", full_name="Super Admin", role="superadmin", is_active=True)
        sa.set_password("super2025")
        db.session.add(sa)
        db.session.commit()
        print("✅ Super Admin dibuat: username=superadmin | password=super2025")

# ===================== RBAC / AUTH =====================
def _admin_allowed(endpoint: str) -> bool:
    if not endpoint:
        return False
    always_ok = {"static", "login", "logout", "dashboard"}
    if endpoint in always_ok:
        return True
    allowed_prefixes = (
        "target_", "target_tagihan_",
        "call_", "noo_", "noo_only_", "ca_", "fl_ao_",
        "rekap", "insentif", "override"
    )
    return any(endpoint.startswith(p) for p in allowed_prefixes)

@app.before_request
def rbac_gate():
    ep = (request.endpoint or "")

    # Izinkan /static dan /logout apa pun kondisinya
    if ep in ("static", "logout"):
        return

    u = session.get("user")

    # Jika SUDAH login lalu buka /login -> kirim ke dashboard
    if ep == "login" and u and not isinstance(u, str):
        return redirect(url_for("dashboard"))

    # Belum login dan bukan /login -> ke /login
    if not u:
        if ep != "login":
            return redirect(url_for("login"))
        return  # sedang di halaman login

    # Session lama masih string -> reset dan minta login
    if isinstance(u, str):
        session.pop("user", None)
        if ep != "login":
            flash("Sesi lama sudah kedaluwarsa, silakan login ulang.", "warning")
            return redirect(url_for("login"))
        return

    role = (u.get("role") or "").lower()

    if role == "superadmin":
        return  # full akses

    if role == "admin":
        if ep == "login":
            return redirect(url_for("dashboard"))
        if _admin_allowed(ep):
            return
        flash("Akses dibatasi untuk Admin. Hubungi Super Admin jika butuh izin.", "danger")
        return redirect(url_for("dashboard"))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            session["user"] = {
                "id": user.id,
                "username": user.username,
                "full_name": user.full_name,
                "role": user.role.lower(),
            }
            flash(f"Selamat datang, {user.full_name}!", "success")
            return redirect(url_for("dashboard"))
        flash("Username atau password salah.", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Anda telah logout.", "info")
    return redirect(url_for("login"))




# ===================== HELPERS =====================
def valid_bulan(s: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", (s or "").strip()))

from typing import Any

def to_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default

def to_float(x: Any, default: float = 0.0) -> float:
    try:
        return float(x)
    except Exception:
        return default


@lru_cache(maxsize=256)
def get_rate(jenis: str, divisi: str=None, kategori: str=None) -> int:
    q = InsentifRate.query.filter_by(jenis=jenis, is_active=True)
    if divisi is not None:
        q = q.filter_by(divisi=divisi)
    else:
        q = q.filter(InsentifRate.divisi.is_(None))
    if kategori is not None:
        q = q.filter_by(kategori=kategori)
    else:
        q = q.filter(InsentifRate.kategori.is_(None))
    r = q.first()
    return int(r.amount) if r else 0

@lru_cache(maxsize=1024)
def get_rate_for(salesman_id: int, jenis: str, bulan: str=None, divisi: str=None, kategori: str=None) -> int:
    # 1) Override spesifik bulan
    q = InsentifOverride.query.filter_by(salesman_id=salesman_id, jenis=jenis, is_active=True)
    if bulan:
        r = q.filter_by(bulan=bulan).first()
        if r:
            return int(r.amount)
    # 2) Override tanpa bulan
    r = q.filter(InsentifOverride.bulan.is_(None)).first()
    if r:
        return int(r.amount)
    # 3) Rate umum
    return get_rate(jenis, divisi, kategori)

def clear_rate_caches():
    get_rate.cache_clear()
    get_rate_for.cache_clear()

def get_salesman_map():
    # returns {id: (name, area, divisi, kategori)}
    rows = Salesman.query.all()
    return {s.id: (s.name, s.area, s.divisi, s.kategori) for s in rows}

def get_password_awal_pair(salesman_id: int, bulan: str):
    """Return (target, actual) untuk PasswordAwal; kalau tidak ada -> (0,0)."""
    r = PasswordAwal.query.filter_by(salesman_id=salesman_id, bulan=bulan).first()
    if not r:
        return 0.0, 0.0
    return float(r.target_amount or 0), float(r.actual_amount or 0)

def calc_password_awal_pass(salesman_id: int, bulan: str, divisi: str) -> bool:
    """
    Password Awal: harus mencapai persentase (default 80%) dari target_awal.
    Persentase bisa diubah di /thresholds jenis 'password_awal' per divisi.
    """
    thresh = get_threshold_percent("password_awal", divisi, default=0.8)
    t, a = get_password_awal_pair(salesman_id, bulan)
    if t <= 0:
        return False
    return (a / t) >= thresh
def get_threshold_percent(kind: str, divisi: str, default=0.8) -> float:
    """
    Ambil persentase threshold untuk (kind, divisi).
    Menerima input 0..1 atau 0..100 (akan dinormalisasi ke 0..1).
    """
    row = Threshold.query.filter_by(kind=kind, divisi=divisi).first()
    if not row or row.percent is None:
        return default
    try:
        p = float(row.percent)
        # Normalisasi bila disimpan dalam skala 0..100
        if p > 1:
            p = p / 100.0
        # Clamp 0..1
        p = max(0.0, min(1.0, p))
        return p
    except Exception:
        return default


# ===================== BUSINESS LOGIC =====================
def calc_pass_penjualan(divisi: str, target_amount: float, actual_amount: float) -> bool:
    if not target_amount:
        return False
    p = get_threshold_percent("penjualan", divisi, default=0.8)
    return (actual_amount or 0) / target_amount >= p

def calc_pass_tagihan(divisi: str, target_amount: float, actual_amount: float) -> bool:
    if not target_amount:
        return False
    p = get_threshold_percent("tagihan", divisi, default=0.8)
    return (actual_amount or 0) / target_amount >= p

def calc_insentif_penjualan(salesman_id: int, bulan: str, divisi: str, kategori: str,
                            target_amount: float, actual_amount: float) -> int:
    # 1) Password Awal wajib lolos
    if not calc_password_awal_pass(salesman_id, bulan, divisi):
        return 0
    # 2) Harus memenuhi threshold penjualan (mengacu ke tabel thresholds; 0..1 atau 0..100 sudah dinormalisasi)
    if not calc_pass_penjualan(divisi, target_amount, actual_amount):
        return 0
    # 3) Lolos → bayar sesuai rate (rate umum/override)
    return get_rate_for(salesman_id, "penjualan", bulan, divisi, kategori)


def calc_insentif_tagihan(salesman_id: int, bulan: str, divisi: str, kategori: str,
                          target_amount: float, actual_amount: float) -> int:
    if not calc_password_awal_pass(salesman_id, bulan, divisi):
        return 0
    if not calc_pass_tagihan(divisi, target_amount, actual_amount):
        return 0
    return get_rate_for(salesman_id, "tagihan", bulan, divisi, kategori)


def calc_insentif_call(salesman_id: int, bulan: str, target_call: int, actual_call: int) -> int:
    """
    Jika actual >= 85% target → insentif = target * rate_call
    Jika actual < 85% target → insentif = actual * rate_call
    rate_call diambil dari InsentifRate / Override
    """
    rate = max(0, get_rate_for(salesman_id, "call", bulan))
    t = max(0, target_call or 0)
    a = max(0, actual_call or 0)
    if t == 0:
        return int(a * rate)
    achievement = a / t
    qty = t if achievement >= 0.85 else a
    return int(qty * rate)

def calc_insentif_noo(salesman_id: int, bulan: str, divisi: str,
                      target_ca: int, actual_ca: int, jumlah_toko_noo: int,
                      target_fl: int, actual_fl: int, target_ao: int, actual_ao: int) -> int:
    total = 0
    # CA: if >=80% → rate 'ca'
    if target_ca and target_ca > 0 and (actual_ca or 0) / target_ca >= 0.8:
        total += get_rate_for(salesman_id, "ca", bulan)
    # NOO/LOST: per toko
    total += (jumlah_toko_noo or 0) * get_rate_for(salesman_id, "noo_item", bulan)
    # FL (MP): if >=80%
    if divisi == "MP" and (target_fl or 0) > 0 and (actual_fl or 0) / target_fl >= 0.8:
        total += get_rate_for(salesman_id, "fl", bulan)
    # AO (AVIAN): if 100%
    if divisi == "AVIAN" and (target_ao or 0) > 0 and (actual_ao or 0) == target_ao:
        total += get_rate_for(salesman_id, "ao", bulan)
    return int(total)

# ===================== ROUTES =====================
@app.route("/")
def root():
    return redirect(url_for("dashboard"))

@app.route("/dashboard")
def dashboard():
    # 1) Total salesman (master)
    total_salesman = Salesman.query.count()

    # 2) PASSWORD AWAL — bulan terakhir & hitung lolos/tidak
    pw_bulan = latest_bulan(PasswordAwal)
    pw_lolos = pw_tidak = 0
    if pw_bulan:
        pw_rows = db.session.query(PasswordAwal, Salesman)\
            .join(Salesman, PasswordAwal.salesman_id == Salesman.id)\
            .filter(PasswordAwal.bulan == pw_bulan).all()
        for r, s in pw_rows:
            if calc_password_awal_pass(s.id, r.bulan, s.divisi):  # pakai threshold di tabel thresholds
                pw_lolos += 1
            else:
                pw_tidak += 1

    # 3) TARGET PENJUALAN — bulan terakhir & hitung lolos/tidak
    pj_bulan = latest_bulan(TargetPenjualan)
    pj_lolos = pj_tidak = 0
    if pj_bulan:
        pj_rows = db.session.query(TargetPenjualan, Salesman)\
            .join(Salesman, TargetPenjualan.salesman_id == Salesman.id)\
            .filter(TargetPenjualan.bulan == pj_bulan).all()
        for r, s in pj_rows:
            if calc_pass_penjualan(s.divisi, r.target_amount, r.actual_amount):  # threshold jenis "penjualan"
                pj_lolos += 1
            else:
                pj_tidak += 1

    # 4) TARGET TAGIHAN — bulan terakhir & hitung lolos/tidak
    tg_bulan = latest_bulan(TargetTagihan)
    tg_lolos = tg_tidak = 0
    if tg_bulan:
        tg_rows = db.session.query(TargetTagihan, Salesman)\
            .join(Salesman, TargetTagihan.salesman_id == Salesman.id)\
            .filter(TargetTagihan.bulan == tg_bulan).all()
        for r, s in tg_rows:
            if calc_pass_tagihan(s.divisi, r.target_amount, r.actual_amount):  # threshold jenis "tagihan"
                tg_lolos += 1
            else:
                tg_tidak += 1

    # 5) NOO ONLY — total jumlah toko NOO (Aktivasi) pada bulan terakhir
    #    (bukan hitung salesman, tapi total angka "jumlah_toko_noo" di bulan terakhir)
    noo_bulan = latest_bulan(NOO)
    noo_total = 0
    if noo_bulan:
        noo_total = db.session.query(func.coalesce(func.sum(NOO.jumlah_toko_noo), 0))\
            .filter(NOO.bulan == noo_bulan).scalar() or 0

    return render_template(
        "dashboard.html",
        total_salesman=total_salesman,
        # password awal
        pw_bulan=pw_bulan, pw_lolos=pw_lolos, pw_tidak=pw_tidak,
        # penjualan
        pj_bulan=pj_bulan, pj_lolos=pj_lolos, pj_tidak=pj_tidak,
        # tagihan
        tg_bulan=tg_bulan, tg_lolos=tg_lolos, tg_tidak=tg_tidak,
        # noo
        noo_bulan=noo_bulan, noo_total=noo_total
    )

# ----- Salesman CRUD -----
@app.route("/salesman")
def salesman_list():
    q = Salesman.query.order_by(Salesman.name.asc()).all()
    return render_template("salesman_list.html", rows=q)

@app.route("/salesman/new", methods=["GET","POST"])
def salesman_new():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        area = request.form.get("area","").strip()
        divisi = request.form.get("divisi","").strip()
        kategori = request.form.get("kategori","").strip()
        if not name or divisi not in ("MP","AVIAN") or kategori not in ("Bronze","Silver","Gold","Platinum"):
            flash("❌ Data tidak valid", "danger")
            return redirect(url_for("salesman_new"))
        db.session.add(Salesman(name=name, area=area, divisi=divisi, kategori=kategori))
        db.session.commit()
        flash("✅ Salesman berhasil ditambahkan", "success")
        return redirect(url_for("salesman_list"))
    return render_template("salesman_form.html", row=None)

@app.route("/salesman/<int:id>/edit", methods=["GET","POST"])
def salesman_edit(id):
    row = Salesman.query.get_or_404(id)
    if request.method == "POST":
        row.name = request.form.get("name","").strip()
        row.area = request.form.get("area","").strip()
        row.divisi = request.form.get("divisi","").strip()
        row.kategori = request.form.get("kategori","").strip()
        db.session.commit()
        flash("✅ Salesman diperbarui", "success")
        return redirect(url_for("salesman_list"))
    return render_template("salesman_form.html", row=row)

@app.route("/salesman/<int:id>/delete", methods=["POST"])
def salesman_delete(id):
    row = Salesman.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ Salesman dihapus", "success")
    return redirect(url_for("salesman_list"))

# ----- TargetPenjualan CRUD -----
@app.route("/target")
def target_list():
    rows = db.session.query(TargetPenjualan, Salesman).join(Salesman, TargetPenjualan.salesman_id == Salesman.id).all()
    return render_template("target_list.html", rows=rows, calc=calc_insentif_penjualan, calc_pass=calc_password_awal_pass)

@app.route("/target/new", methods=["GET","POST"])
def target_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan", "").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM (contoh 2025-10).", "danger")
            return redirect(url_for("target_new"))

        salesman_id = int(request.form.get("salesman_id"))
        target_amount = to_float(request.form.get("target_amount"), 0)
        actual_amount = to_float(request.form.get("actual_amount"), 0)
        if target_amount < 0 or actual_amount < 0:
            flash("❌ Target/Aktual tidak boleh negatif.", "danger")
            return redirect(url_for("target_new"))

        db.session.add(TargetPenjualan(
            salesman_id=salesman_id,
            bulan=bulan,
            target_amount=target_amount,
            actual_amount=actual_amount
        ))
        db.session.commit()
        flash("✅ Target Penjualan tersimpan", "success")
        return redirect(url_for("target_list"))

    return render_template("target_form.html", salesmen=salesmen)

@app.route("/target/<int:id>/delete", methods=["POST"])
def target_delete(id):
    row = TargetPenjualan.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ Target Penjualan dihapus", "success")
    return redirect(url_for("target_list"))

# ----- TargetTagihan CRUD -----
@app.route("/target_tagihan")
def target_tagihan_list():
    rows = db.session.query(TargetTagihan, Salesman).join(Salesman, TargetTagihan.salesman_id == Salesman.id).all()
    return render_template("target_tagihan_list.html", rows=rows, calc=calc_insentif_tagihan, calc_pass=calc_password_awal_pass)

@app.route("/target_tagihan/new", methods=["GET","POST"])
def target_tagihan_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan", "").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM (contoh 2025-10).", "danger")
            return redirect(url_for("target_tagihan_new"))

        salesman_id = int(request.form.get("salesman_id"))
        target_amount = to_float(request.form.get("target_amount"), 0)
        actual_amount = to_float(request.form.get("actual_amount"), 0)
        if target_amount < 0 or actual_amount < 0:
            flash("❌ Target/Aktual tidak boleh negatif.", "danger")
            return redirect(url_for("target_tagihan_new"))

        db.session.add(TargetTagihan(
            salesman_id=salesman_id,
            bulan=bulan,
            target_amount=target_amount,
            actual_amount=actual_amount
        ))
        db.session.commit()
        flash("✅ Target Tagihan tersimpan", "success")
        return redirect(url_for("target_tagihan_list"))

    return render_template("target_tagihan_form.html", salesmen=salesmen)

@app.route("/target_tagihan/<int:id>/delete", methods=["POST"])
def target_tagihan_delete(id):
    row = TargetTagihan.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ Target Tagihan dihapus", "success")
    return redirect(url_for("target_tagihan_list"))

# ----- NOO CRUD -----
@app.route("/noo")
def noo_list():
    rows = db.session.query(NOO, Salesman).join(Salesman, NOO.salesman_id == Salesman.id).all()
    return render_template("noo_list.html", rows=rows, calc_noo=calc_insentif_noo)

@app.route("/noo/new", methods=["GET","POST"])
def noo_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        data = {k: request.form.get(k) for k in request.form.keys()}
        def _to_int(x):
            try:
                return int(x)
            except:
                return 0
        row = NOO(
            salesman_id = int(data.get("salesman_id")),
            bulan = data.get("bulan"),
            target_ca = _to_int(data.get("target_ca")),
            actual_ca = _to_int(data.get("actual_ca")),
            jumlah_toko_noo = _to_int(data.get("jumlah_toko_noo")),
            target_fl = _to_int(data.get("target_fl")),
            actual_fl = _to_int(data.get("actual_fl")),
            target_ao = _to_int(data.get("target_ao")),
            actual_ao = _to_int(data.get("actual_ao")),
        )
        db.session.add(row)
        db.session.commit()
        flash("✅ Data NOO tersimpan", "success")
        return redirect(url_for("noo_list"))
    return render_template("noo_form.html", salesmen=salesmen)

@app.route("/noo/<int:id>/edit", methods=["GET","POST"])
def noo_edit(id):
    row = NOO.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        data = {k: request.form.get(k) for k in request.form.keys()}
        def _to_int(x):
            try:
                return int(x)
            except:
                return 0
        row.salesman_id = int(data.get("salesman_id"))
        row.bulan = data.get("bulan")
        row.target_ca = _to_int(data.get("target_ca"))
        row.actual_ca = _to_int(data.get("actual_ca"))
        row.jumlah_toko_noo = _to_int(data.get("jumlah_toko_noo"))
        row.target_fl = _to_int(data.get("target_fl"))
        row.actual_fl = _to_int(data.get("actual_fl"))
        row.target_ao = _to_int(data.get("target_ao"))
        row.actual_ao = _to_int(data.get("actual_ao"))
        db.session.commit()
        flash("✅ Data NOO diperbarui", "success")
        return redirect(url_for("noo_list"))
    return render_template("noo_form.html", row=row, salesmen=salesmen)

@app.route("/noo/<int:id>/delete", methods=["POST"])
def noo_delete(id):
    row = NOO.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ Data NOO dihapus", "success")
    return redirect(url_for("noo_list"))

# ===================== CA ONLY =====================
@app.route("/ca")
def ca_list():
    rows = db.session.query(NOO, Salesman)\
        .join(Salesman, NOO.salesman_id == Salesman.id)\
        .filter( (NOO.target_ca > 0) | (NOO.actual_ca > 0) ).all()
    return render_template("ca_list.html", rows=rows)

@app.route("/ca/new", methods=["GET","POST"])
def ca_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("ca_new"))

        salesman_id = int(request.form.get("salesman_id"))
        tca = to_int(request.form.get("target_ca"), 0)
        aca = to_int(request.form.get("actual_ca"), 0)
        if tca < 0 or aca < 0:
            flash("❌ Target/Actual CA tidak boleh negatif.", "danger")
            return redirect(url_for("ca_new"))

        row = NOO(
            salesman_id=salesman_id, bulan=bulan,
            target_ca=tca, actual_ca=aca,
            jumlah_toko_noo=0, target_fl=0, actual_fl=0,
            target_ao=0, actual_ao=0
        )
        db.session.add(row); db.session.commit()
        flash("✅ CA tersimpan", "success")
        return redirect(url_for("ca_list"))
    return render_template("ca_form.html", salesmen=salesmen)

@app.route("/ca/<int:id>/edit", methods=["GET","POST"])
def ca_edit(id):
    row = NOO.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("ca_edit", id=id))

        row.salesman_id = int(request.form.get("salesman_id"))
        row.bulan = bulan
        row.target_ca = to_int(request.form.get("target_ca"), 0)
        row.actual_ca = to_int(request.form.get("actual_ca"), 0)
        if row.target_ca < 0 or row.actual_ca < 0:
            flash("❌ Target/Actual CA tidak boleh negatif.", "danger")
            return redirect(url_for("ca_edit", id=id))

        # tetap CA only
        row.jumlah_toko_noo = 0
        row.target_fl = row.actual_fl = 0
        row.target_ao = row.actual_ao = 0

        db.session.commit()
        flash("✅ CA diperbarui", "success")
        return redirect(url_for("ca_list"))
    return render_template("ca_form.html", row=row, salesmen=salesmen)

@app.route("/ca/<int:id>/delete", methods=["POST"])
def ca_delete(id):
    row = NOO.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ CA dihapus", "success")
    return redirect(url_for("ca_list"))

# ===================== NOO-ONLY & FL/AO-ONLY =====================
@app.route("/noo-only")
def noo_only_list():
    rows = db.session.query(NOO, Salesman)\
        .join(Salesman, NOO.salesman_id == Salesman.id)\
        .filter(NOO.jumlah_toko_noo > 0).all()
    return render_template("noo_only_list.html", rows=rows)

@app.route("/noo-only/new", methods=["GET","POST"])
def noo_only_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("noo_only_new"))

        salesman_id = int(request.form.get("salesman_id"))
        j = to_int(request.form.get("jumlah_toko_noo"), 0)
        if j < 0:
            flash("❌ Jumlah toko NOO/LOST tidak boleh negatif.", "danger")
            return redirect(url_for("noo_only_new"))

        row = NOO(
            salesman_id=salesman_id, bulan=bulan,
            target_ca=0, actual_ca=0,
            jumlah_toko_noo=j,
            target_fl=0, actual_fl=0,
            target_ao=0, actual_ao=0
        )
        db.session.add(row); db.session.commit()
        flash("✅ NOO/LOST tersimpan", "success")
        return redirect(url_for("noo_only_list"))
    return render_template("noo_only_form.html", salesmen=salesmen)

@app.route("/noo-only/<int:id>/edit", methods=["GET","POST"])
def noo_only_edit(id):
    row = NOO.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("noo_only_edit", id=id))

        row.salesman_id = int(request.form.get("salesman_id"))
        row.bulan = bulan
        row.jumlah_toko_noo = to_int(request.form.get("jumlah_toko_noo"), 0)
        if row.jumlah_toko_noo < 0:
            flash("❌ Jumlah toko NOO/LOST tidak boleh negatif.", "danger")
            return redirect(url_for("noo_only_edit", id=id))

        # tetap NOO only
        row.target_ca = row.actual_ca = 0
        row.target_fl = row.actual_fl = 0
        row.target_ao = row.actual_ao = 0

        db.session.commit()
        flash("✅ NOO/LOST diperbarui", "success")
        return redirect(url_for("noo_only_list"))
    return render_template("noo_only_form.html", row=row, salesmen=salesmen)

@app.route("/noo-only/<int:id>/delete", methods=["POST"])
def noo_only_delete(id):
    row = NOO.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ NOO/LOST dihapus", "success")
    return redirect(url_for("noo_only_list"))

@app.route("/fl-ao")
def fl_ao_list():
    rows = db.session.query(NOO, Salesman)\
        .join(Salesman, NOO.salesman_id == Salesman.id)\
        .filter( (NOO.target_fl > 0) | (NOO.actual_fl > 0) | (NOO.target_ao > 0) | (NOO.actual_ao > 0) ).all()
    return render_template("fl_ao_list.html", rows=rows)

@app.route("/fl-ao/new", methods=["GET","POST"])
def fl_ao_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("fl_ao_new"))

        sid = int(request.form.get("salesman_id"))
        s = Salesman.query.get(sid)
        if not s:
            flash("❌ Salesman tidak ditemukan.", "danger")
            return redirect(url_for("fl_ao_new"))

        if s.divisi == "MP":
            tfl = to_int(request.form.get("target_fl"), 0)
            afl = to_int(request.form.get("actual_fl"), 0)
            if tfl < 0 or afl < 0:
                flash("❌ FL tidak boleh negatif.", "danger")
                return redirect(url_for("fl_ao_new"))
            tao = aao = 0
        else:  # AVIAN
            tao = to_int(request.form.get("target_ao"), 0)
            aao = to_int(request.form.get("actual_ao"), 0)
            if tao < 0 or aao < 0:
                flash("❌ AO tidak boleh negatif.", "danger")
                return redirect(url_for("fl_ao_new"))
            tfl = afl = 0

        row = NOO(
            salesman_id=sid, bulan=bulan,
            target_ca=0, actual_ca=0, jumlah_toko_noo=0,
            target_fl=tfl, actual_fl=afl,
            target_ao=tao, actual_ao=aao
        )
        db.session.add(row); db.session.commit()
        flash("✅ FL/AO tersimpan", "success")
        return redirect(url_for("fl_ao_list"))
    return render_template("fl_ao_form.html", salesmen=salesmen)

@app.route("/fl-ao/<int:id>/edit", methods=["GET","POST"])
def fl_ao_edit(id):
    row = NOO.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("fl_ao_edit", id=id))

        sid = int(request.form.get("salesman_id"))
        s = Salesman.query.get(sid)
        if not s:
            flash("❌ Salesman tidak ditemukan.", "danger")
            return redirect(url_for("fl_ao_edit", id=id))

        row.salesman_id = sid
        row.bulan = bulan

        if s.divisi == "MP":
            row.target_fl = to_int(request.form.get("target_fl"), 0)
            row.actual_fl = to_int(request.form.get("actual_fl"), 0)
            if row.target_fl < 0 or row.actual_fl < 0:
                flash("❌ FL tidak boleh negatif.", "danger")
                return redirect(url_for("fl_ao_edit", id=id))
            row.target_ao = row.actual_ao = 0
        else:
            row.target_ao = to_int(request.form.get("target_ao"), 0)
            row.actual_ao = to_int(request.form.get("actual_ao"), 0)
            if row.target_ao < 0 or row.actual_ao < 0:
                flash("❌ AO tidak boleh negatif.", "danger")
                return redirect(url_for("fl_ao_edit", id=id))
            row.target_fl = row.actual_fl = 0

        # tetap FL/AO only
        row.target_ca = row.actual_ca = 0
        row.jumlah_toko_noo = 0

        db.session.commit()
        flash("✅ FL/AO diperbarui", "success")
        return redirect(url_for("fl_ao_list"))
    return render_template("fl_ao_form.html", row=row, salesmen=salesmen)

@app.route("/fl-ao/<int:id>/delete", methods=["POST"])
def fl_ao_delete(id):
    row = NOO.query.get_or_404(id)
    db.session.delete(row)
    db.session.commit()
    flash("✅ FL/AO dihapus", "success")
    return redirect(url_for("fl_ao_list"))


# ===================== PASSWORD AWAL =====================
@app.route("/password-awal")
def password_awal_list():
    rows = db.session.query(PasswordAwal, Salesman)\
        .join(Salesman, PasswordAwal.salesman_id == Salesman.id)\
        .order_by(PasswordAwal.bulan.desc(), Salesman.name.asc()).all()
    try:
        return render_template("password_awal_list.html", rows=rows, pass_check=calc_password_awal_pass)
    except:
        # fallback sederhana
        return render_template_string("""
        <div class="container p-4">
          <h3>Target Awal (Password)</h3>
          <p><a href="{{ url_for('password_awal_new') }}">Tambah</a></p>
          <table border="1" cellpadding="6" cellspacing="0">
            <tr>
              <th>Bulan</th><th>Nama Sales</th><th>Divisi</th>
              <th>Target Awal</th><th>Aktual</th><th>Lolos?</th><th>Aksi</th>
            </tr>
            {% for r,s in rows %}
            <tr>
              <td>{{ r.bulan }}</td>
              <td>{{ s.name }}</td>
              <td>{{ s.divisi }}</td>
              <td>{{ "{:,.0f}".format(r.target_amount or 0) }}</td>
              <td>{{ "{:,.0f}".format(r.actual_amount or 0) }}</td>
              <td>{{ 'Ya' if pass_check(s.id, r.bulan, s.divisi) else 'Tidak' }}</td>
              <td>
                <a href="{{ url_for('password_awal_edit', id=r.id) }}">Edit</a> |
                <form action="{{ url_for('password_awal_delete', id=r.id) }}" method="post" style="display:inline" onsubmit="return confirm('Hapus data ini?')">
                  <button type="submit">Delete</button>
                </form>
              </td>
            </tr>
            {% endfor %}
          </table>
        </div>
        """, rows=rows, pass_check=calc_password_awal_pass)

@app.route("/password-awal/new", methods=["GET","POST"])
def password_awal_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = (request.form.get("bulan") or "").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("password_awal_new"))
        sid = int(request.form.get("salesman_id"))
        t = to_float(request.form.get("target_amount"), 0.0)
        a = to_float(request.form.get("actual_amount"), 0.0)
        if t < 0 or a < 0:
            flash("❌ Target/Aktual tidak boleh negatif.", "danger")
            return redirect(url_for("password_awal_new"))
        row = PasswordAwal(salesman_id=sid, bulan=bulan, target_amount=t, actual_amount=a)
        try:
            db.session.add(row); db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash("❌ Duplikat data untuk salesman & bulan yang sama.", "danger")
            return redirect(url_for("password_awal_new"))
        flash("✅ Target Awal tersimpan", "success")
        return redirect(url_for("password_awal_list"))
    try:
        return render_template("password_awal_form.html", salesmen=salesmen, row=None)
    except:
        return render_template_string("""
        <div class="container p-4">
          <h3>Tambah Target Awal</h3>
          <form method="post">
            <label>Bulan (YYYY-MM)</label><br>
            <input type="text" name="bulan" placeholder="2025-10" required><br><br>

            <label>Salesman</label><br>
            <select name="salesman_id" required>
              {% for s in salesmen %}
                <option value="{{ s.id }}">{{ s.name }} ({{ s.divisi }})</option>
              {% endfor %}
            </select><br><br>

            <label>Target Awal</label><br>
            <input type="number" step="0.01" name="target_amount" required><br><br>

            <label>Aktual</label><br>
            <input type="number" step="0.01" name="actual_amount" value="0"><br><br>

            <button type="submit">Simpan</button>
            <a href="{{ url_for('password_awal_list') }}">Batal</a>
          </form>
        </div>
        """, salesmen=salesmen, row=None)

@app.route("/password-awal/<int:id>/edit", methods=["GET","POST"])
def password_awal_edit(id):
    row = PasswordAwal.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = (request.form.get("bulan") or "").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM.", "danger")
            return redirect(url_for("password_awal_edit", id=id))
        row.salesman_id = int(request.form.get("salesman_id"))
        row.bulan = bulan
        row.target_amount = to_float(request.form.get("target_amount"), 0.0)
        row.actual_amount = to_float(request.form.get("actual_amount"), 0.0)
        if row.target_amount < 0 or row.actual_amount < 0:
            flash("❌ Target/Aktual tidak boleh negatif.", "danger")
            return redirect(url_for("password_awal_edit", id=id))
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("❌ Duplikat data untuk salesman & bulan yang sama.", "danger")
            return redirect(url_for("password_awal_edit", id=id))
        flash("✅ Target Awal diperbarui", "success")
        return redirect(url_for("password_awal_list"))
    try:
        return render_template("password_awal_form.html", salesmen=salesmen, row=row)
    except:
        return render_template_string("""
        <div class="container p-4">
          <h3>Edit Target Awal</h3>
          <form method="post">
            <label>Bulan (YYYY-MM)</label><br>
            <input type="text" name="bulan" value="{{ row.bulan }}" required><br><br>

            <label>Salesman</label><br>
            <select name="salesman_id" required>
              {% for s in salesmen %}
                <option value="{{ s.id }}" {{ 'selected' if s.id==row.salesman_id else '' }}>
                  {{ s.name }} ({{ s.divisi }})
                </option>
              {% endfor %}
            </select><br><br>

            <label>Target Awal</label><br>
            <input type="number" step="0.01" name="target_amount" value="{{ row.target_amount }}"><br><br>

            <label>Aktual</label><br>
            <input type="number" step="0.01" name="actual_amount" value="{{ row.actual_amount }}"><br><br>

            <button type="submit">Simpan</button>
            <a href="{{ url_for('password_awal_list') }}">Batal</a>
          </form>
        </div>
        """, salesmen=salesmen, row=row)

@app.route("/password-awal/<int:id>/delete", methods=["POST"])
def password_awal_delete(id):
    row = PasswordAwal.query.get_or_404(id)
    db.session.delete(row); db.session.commit()
    flash("✅ Target Awal dihapus", "success")
    return redirect(url_for("password_awal_list"))


# ===================== CALL CRUD =====================
@app.route("/call")
def call_list():
    rows = db.session.query(Call, Salesman)\
        .join(Salesman, Call.salesman_id == Salesman.id)\
        .order_by(Call.bulan.desc(), Salesman.name.asc())\
        .all()
    return render_template("call_list.html", rows=rows, calc_call=calc_insentif_call)

@app.route("/call/new", methods=["GET","POST"])
def call_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM (contoh 2025-10).", "danger")
            return redirect(url_for("call_new"))

        try:
            row = Call(
                salesman_id = int(request.form["salesman_id"]),
                bulan        = bulan,
                target_call  = to_int(request.form.get("target_call"), 0),
                actual_call  = to_int(request.form.get("actual_call"), 0),
            )
        except Exception:
            flash("❌ Data tidak valid.", "danger")
            return redirect(url_for("call_new"))

        if row.target_call < 0 or row.actual_call < 0:
            flash("❌ Target/Actual Call tidak boleh negatif.", "danger")
            return redirect(url_for("call_new"))

        db.session.add(row); db.session.commit()
        flash("✅ Data Call tersimpan", "success")
        return redirect(url_for("call_list"))
    return render_template("call_form.html", salesmen=salesmen, row=None)

@app.route("/call/<int:id>/edit", methods=["GET","POST"])
def call_edit(id):
    row = Call.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method == "POST":
        bulan = request.form.get("bulan","").strip()
        if not valid_bulan(bulan):
            flash("❌ Format bulan wajib YYYY-MM (contoh 2025-10).", "danger")
            return redirect(url_for("call_edit", id=id))

        row.salesman_id = int(request.form["salesman_id"])
        row.bulan       = bulan
        row.target_call = to_int(request.form.get("target_call"), 0)
        row.actual_call = to_int(request.form.get("actual_call"), 0)

        if row.target_call < 0 or row.actual_call < 0:
            flash("❌ Target/Actual Call tidak boleh negatif.", "danger")
            return redirect(url_for("call_edit", id=id))

        db.session.commit()
        flash("✅ Data Call diperbarui", "success")
        return redirect(url_for("call_list"))
    return render_template("call_form.html", salesmen=salesmen, row=row)

@app.route("/call/<int:id>/delete", methods=["POST"])
def call_delete(id):
    row = Call.query.get_or_404(id)
    db.session.delete(row); db.session.commit()
    flash("✅ Data Call dihapus", "success")
    return redirect(url_for("call_list"))

# ----- Rekap -----
@app.route("/rekap")
def rekap():
    bulan = request.args.get("bulan", "").strip()
    smap = get_salesman_map()

    # Queries terpisah
    penjualan_q = TargetPenjualan.query
    tagihan_q   = TargetTagihan.query
    noo_q       = NOO.query
    call_q      = Call.query
    if bulan:
        penjualan_q = penjualan_q.filter(TargetPenjualan.bulan == bulan)
        tagihan_q   = tagihan_q.filter(TargetTagihan.bulan == bulan)
        noo_q       = noo_q.filter(NOO.bulan == bulan)
        call_q      = call_q.filter(Call.bulan == bulan)

    penjualan_rows = penjualan_q.all()
    tagihan_rows   = tagihan_q.all()
    noo_rows       = noo_q.all()
    call_rows      = call_q.all()

    from collections import defaultdict
    penjualan_map = defaultdict(int)
    tagihan_map   = defaultdict(int)
    noo_map       = defaultdict(int)
    call_map      = defaultdict(int)

    for r in penjualan_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        penjualan_map[(bkey, r.salesman_id)] += calc_insentif_penjualan(
            r.salesman_id, bkey, divisi, kategori, r.target_amount, r.actual_amount
        )

    for r in tagihan_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        tagihan_map[(bkey, r.salesman_id)] += calc_insentif_tagihan(
            r.salesman_id, bkey, divisi, kategori, r.target_amount, r.actual_amount
        )

    for r in noo_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        noo_map[(bkey, r.salesman_id)] += calc_insentif_noo(
            r.salesman_id, bkey, divisi, r.target_ca, r.actual_ca,
            r.jumlah_toko_noo, r.target_fl, r.actual_fl, r.target_ao, r.actual_ao
        )

    for r in call_rows:
        bkey = bulan or r.bulan
        # Jika ingin dihitung, ganti 0 dengan calc_insentif_call(...)
        call_map[(bkey, r.salesman_id)] += 0

    # Compose rows
    ids = list(smap.keys())
    keys = set(list(penjualan_map.keys()) + list(tagihan_map.keys()) + list(noo_map.keys()) + list(call_map.keys()))
    if bulan:
        for sid in ids:
            keys.add((bulan, sid))

    results = []
    for (b, sid) in sorted(keys, key=lambda x: (x[0], smap.get(x[1], ("", "", "", ""))[0])):
        name, area, divisi, kategori = smap.get(sid, ("?","?","?","?"))
        total_penjualan = penjualan_map.get((b, sid), 0)
        total_tagihan   = tagihan_map.get((b, sid), 0)
        total_noo       = noo_map.get((b, sid), 0)
        total_call      = call_map.get((b, sid), 0)
        total = total_penjualan + total_tagihan + total_noo + total_call
        results.append({
            "bulan": b, "salesman_id": sid, "name": name, "area": area,
            "divisi": divisi, "kategori": kategori,
            "total_penjualan": total_penjualan, "total_tagihan": total_tagihan,
            "total_noo": total_noo, "total_call": total_call, "total": total
        })

    return render_template("rekap.html", rows=results, bulan=bulan)

@app.route("/rekap/export")
def rekap_export():
    bulan = request.args.get("bulan", "").strip()
    smap = get_salesman_map()

    penjualan_rows = TargetPenjualan.query
    tagihan_rows   = TargetTagihan.query
    noo_rows       = NOO.query
    call_rows      = Call.query
    if bulan:
        penjualan_rows = penjualan_rows.filter(TargetPenjualan.bulan == bulan)
        tagihan_rows   = tagihan_rows.filter(TargetTagihan.bulan == bulan)
        noo_rows       = noo_rows.filter(NOO.bulan == bulan)
        call_rows      = call_rows.filter(Call.bulan == bulan)

    penjualan_rows = penjualan_rows.all()
    tagihan_rows   = tagihan_rows.all()
    noo_rows       = noo_rows.all()
    call_rows      = call_rows.all()

    from collections import defaultdict
    penjualan_map = defaultdict(int)
    tagihan_map   = defaultdict(int)
    noo_map       = defaultdict(int)
    call_map      = defaultdict(int)

    for r in penjualan_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        penjualan_map[(bkey, r.salesman_id)] += calc_insentif_penjualan(
            r.salesman_id, bkey, divisi, kategori, r.target_amount, r.actual_amount
        )
    for r in tagihan_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        tagihan_map[(bkey, r.salesman_id)] += calc_insentif_tagihan(
            r.salesman_id, bkey, divisi, kategori, r.target_amount, r.actual_amount
        )
    for r in noo_rows:
        name, area, divisi, kategori = smap.get(r.salesman_id, ("?","?","MP","Bronze"))
        bkey = bulan or r.bulan
        noo_map[(bkey, r.salesman_id)] += calc_insentif_noo(
            r.salesman_id, bkey, divisi, r.target_ca, r.actual_ca,
            r.jumlah_toko_noo, r.target_fl, r.actual_fl, r.target_ao, r.actual_ao
        )
    for r in call_rows:
        bkey = bulan or r.bulan
        # TODO: ganti 0 dengan perhitungan insentif call saat logic final sudah siap
        call_map[(bkey, r.salesman_id)] += 0

    keys = set(list(penjualan_map.keys()) + list(tagihan_map.keys())  + list(noo_map.keys()) + list(call_map.keys()))
    rows = []
    for (b, sid) in sorted(keys):
        name, area, divisi, kategori = smap.get(sid, ("?","?","?","?"))
        total_penjualan = penjualan_map.get((b, sid), 0)
        total_tagihan   = tagihan_map.get((b, sid), 0)
        total_noo       = noo_map.get((b, sid), 0)
        total_call      = call_map.get((b, sid), 0)
        total = total_penjualan + total_tagihan + total_noo + total_call
        rows.append({
            "Bulan": b,
            "Nama Sales": name,
            "Area": area,
            "Divisi": divisi,
            "Kategori": kategori,
            "Insentif Penjualan": total_penjualan,
            "Insentif Tagihan": total_tagihan,
            "Insentif Gabungan (NOO / CA / FL–AO)": total_noo,
            "Insentif Call": total_call,
            "Total Insentif": total
        })

    import pandas as pd, io
    df = pd.DataFrame(rows)

    # ====== Mulai styling Excel ======
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from datetime import datetime

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Rekap"
        start_row = 3  # ruang untuk judul & subjudul
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        # Brand color PTSP
        MAROON = "7A1F1F"
        LIGHT_GRAY = "F2F3F7"
        WHITE = "FFFFFF"
        BORDER_CLR = "E5E7EB"

        # Title & subtitle
        last_col = get_column_letter(ws.max_column)
        ws.merge_cells(f"A1:{last_col}1")
        ws.merge_cells(f"A2:{last_col}2")
        ws["A1"].value = "REKAP INSENTIF SALES"
        ws["A2"].value = f"Periode: {bulan or 'Semua Bulan'} • Dibuat: {datetime.now().strftime('%d %b %Y %H:%M')}"
        ws["A1"].font = Font(bold=True, size=16, color=WHITE)
        ws["A2"].font = Font(size=11, color=WHITE)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill("solid", fgColor=MAROON)
        ws["A2"].fill = PatternFill("solid", fgColor=MAROON)
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 20

        # Header style
        header_row = start_row + 1
        thin = Side(style="thin", color=BORDER_CLR)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.row_dimensions[header_row].height = 22

        # Body style + borders
        for r in range(header_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Auto width
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        # Freeze dan filter
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)  # freeze sampai header
        ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

        # Format Rupiah untuk kolom uang
        rupiah_cols = [
            "Insentif Penjualan",
            "Insentif Tagihan",
            "Insentif Gabungan (NOO / CA / FL–AO)",
            "Insentif Call",
            "Total Insentif",
        ]
        # map nama kolom -> index kolom di sheet
        header_map = {ws.cell(row=header_row, column=i).value: i for i in range(1, ws.max_column + 1)}
        for cname in rupiah_cols:
            if cname not in header_map:
                continue
            cidx = header_map[cname]
            for r in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=r, column=cidx).number_format = '"Rp" #,##0'

        # Data bar untuk Total Insentif
        if "Total Insentif" in header_map:
            cidx = header_map["Total Insentif"]
            col_letter = get_column_letter(cidx)
            data_range = f"{col_letter}{header_row + 1}:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                data_range,
                DataBarRule(start_type="min", end_type="max", showValue="None", color="A0C4FF")
            )

        # Baris TOTAL di bawah
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")

        for cname in rupiah_cols:
            if cname not in header_map:
                continue
            cidx = header_map[cname]
            col_letter = get_column_letter(cidx)
            ws.cell(row=total_row, column=cidx).value = f"=SUM({col_letter}{header_row + 1}:{col_letter}{total_row - 1})"
            ws.cell(row=total_row, column=cidx).number_format = '"Rp" #,##0'
            ws.cell(row=total_row, column=cidx).font = Font(bold=True)

        # Background tipis pada baris total
        for c in range(1, ws.max_column + 1):
            ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="FFF8E1")

    output.seek(0)
    filename = f"Rekap_Insentif_{bulan or 'semua'}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ===================== INSENTIF RATE CRUD =====================
@app.route("/insentif")
def insentif_list():
    rows = InsentifRate.query.order_by(
        InsentifRate.jenis.asc(),
        InsentifRate.divisi.asc().nullsfirst(),
        InsentifRate.kategori.asc().nullsfirst()
    ).all()
    try:
        return render_template("insentif_list.html", rows=rows)
    except:
        return render_template_string("""
        <div class="container p-4">
          <h3>Insentif Rate</h3>
          <p>
            <a href="{{ url_for('insentif_template') }}">Download Template</a> |
            <a href="{{ url_for('insentif_import') }}">Import Excel</a> |
            <a href="{{ url_for('insentif_new') }}">Tambah</a>
            | <a href="{{ url_for('override_list') }}">Override per Sales</a>
          </p>
          <table border="1" cellpadding="6" cellspacing="0">
            <tr><th>Jenis</th><th>Divisi</th><th>Kategori</th><th>Amount</th><th>Active</th><th>Aksi</th></tr>
            {% for r in rows %}
            <tr>
              <td>{{ r.jenis }}</td><td>{{ r.divisi or '-' }}</td><td>{{ r.kategori or '-' }}</td>
              <td>{{ "{:,}".format(r.amount) }}</td><td>{{ 'Ya' if r.is_active else 'Tidak' }}</td>
              <td>
                <a href="{{ url_for('insentif_edit', id=r.id) }}">Edit</a> |
                <form action="{{ url_for('insentif_delete', id=r.id) }}" method="post" style="display:inline" onsubmit="return confirm('Hapus?')">
                  <button type="submit">Delete</button>
                </form>
              </td>
            </tr>
            {% endfor %}
          </table>
        </div>
        """, rows=rows)

@app.route("/insentif/new", methods=["GET","POST"])
def insentif_new():
    if request.method == "POST":
        r = InsentifRate(
            jenis=request.form["jenis"],
            divisi=request.form.get("divisi") or None,
            kategori=request.form.get("kategori") or None,
            amount=int(request.form.get("amount") or 0),
            is_active=bool(request.form.get("is_active"))
        )
        db.session.add(r); db.session.commit()
        clear_rate_caches()
        flash("✅ Insentif ditambahkan","success")
        return redirect(url_for("insentif_list"))

    try:
        return render_template("insentif_form.html", row=None)
    except:
        return render_template_string(_inline_rate_form(), row=None)

@app.route("/insentif/<int:id>/edit", methods=["GET","POST"])
def insentif_edit(id):
    r = InsentifRate.query.get_or_404(id)
    if request.method == "POST":
        r.jenis = request.form["jenis"]
        r.divisi = request.form.get("divisi") or None
        r.kategori = request.form.get("kategori") or None
        r.amount = int(request.form.get("amount") or 0)
        r.is_active = bool(request.form.get("is_active"))
        db.session.commit()
        clear_rate_caches()
        flash("✅ Insentif diperbarui","success")
        return redirect(url_for("insentif_list"))
    try:
        return render_template("insentif_form.html", row=r)
    except:
        return render_template_string(_inline_rate_form(), row=r)

@app.route("/insentif/<int:id>/delete", methods=["POST"])
def insentif_delete(id):
    r = InsentifRate.query.get_or_404(id)
    db.session.delete(r); db.session.commit()
    clear_rate_caches()
    flash("✅ Insentif dihapus","success")
    return redirect(url_for("insentif_list"))

def _inline_rate_form():
    # fallback form sederhana bila template tidak tersedia
    return """
    <div class="container p-4">
      <h3>{{ 'Edit' if row else 'Tambah' }} Insentif Rate</h3>
      <form method="post">
        <label>Jenis</label>
        <select name="jenis" required>
          {% set jenis_opts = ['penjualan','tagihan','ca','noo_item','fl','ao','call'] %}
          {% for j in jenis_opts %}
            <option value="{{j}}" {{ 'selected' if row and row.jenis==j else '' }}>{{ j }}</option>
          {% endfor %}
        </select><br><br>

        <label>Divisi</label>
        <select name="divisi">
          <option value="">—</option>
          <option value="MP" {{ 'selected' if row and row.divisi=='MP' else '' }}>MP</option>
          <option value="AVIAN" {{ 'selected' if row and row.divisi=='AVIAN' else '' }}>AVIAN</option>
        </select><br><br>

        <label>Kategori</label>
        <select name="kategori">
          <option value="">—</option>
          {% for k in ['Bronze','Silver','Gold','Platinum'] %}
            <option value="{{k}}" {{ 'selected' if row and row.kategori==k else '' }}>{{ k }}</option>
          {% endfor %}
        </select><br><br>

        <label>Amount (Rp)</label>
        <input type="number" name="amount" value="{{ row.amount if row else 0 }}" required><br><br>

        <label><input type="checkbox" name="is_active" {{ 'checked' if (row and row.is_active) or (not row) else '' }}> Active</label><br><br>

        <button type="submit">Simpan</button>
        <a href="{{ url_for('insentif_list') }}">Batal</a>
      </form>
    </div>
    """

def _parse_bool(val: str):
    if val is None:
        return None
    v = str(val).strip().lower()
    if v in ("1", "true", "ya", "y", "aktif", "active", "on"):
        return True
    if v in ("0", "false", "tidak", "n", "nonaktif", "inactive", "off"):
        return False
    return None  # kalau tidak cocok, abaikan sebagai 'tanpa filter'

@app.route("/insentif/export")
def insentif_export():
    # ---- pilih model yang ada ----
    Model = globals().get("InsentifRate")
    if Model is None:
        return abort(500, "Model InsentifRate tidak ditemukan.")

    # ---- ambil filter ----
    jenis_q    = (request.args.get("jenis") or "").strip()
    divisi_q   = (request.args.get("divisi") or "").strip()
    kategori_q = (request.args.get("kategori") or "").strip()
    active_q   = _parse_bool(request.args.get("is_active"))

    q = Model.query
    if jenis_q:
        q = q.filter(Model.jenis == jenis_q)
    if hasattr(Model, "divisi") and divisi_q:
        q = q.filter(Model.divisi == divisi_q)
    if hasattr(Model, "kategori") and kategori_q:
        q = q.filter(Model.kategori == kategori_q)
    if hasattr(Model, "is_active") and active_q is not None:
        q = q.filter(Model.is_active == active_q)

    # ---- order aman ----
    try:
        q = q.order_by(
            asc(Model.jenis),
            asc(Model.divisi).nullsfirst() if hasattr(Model, "divisi") else asc(Model.jenis),
            asc(Model.kategori).nullsfirst() if hasattr(Model, "kategori") else asc(Model.jenis),
        )
    except Exception:
        q = q.order_by(
            Model.jenis.asc(),
            getattr(Model, "divisi", Model.jenis).asc(),
            getattr(Model, "kategori", Model.jenis).asc(),
        )

    rows = q.all()

    # ---- siapkan data untuk Excel ----
    data = []
    for i, r in enumerate(rows, 1):
        data.append({
            "No": i,
            "Jenis": getattr(r, "jenis", "") or "",
            "Divisi": getattr(r, "divisi", "") or "",
            "Kategori": getattr(r, "kategori", "") or "",
            "Nominal (Rp)": int(getattr(r, "amount", 0) or 0),
            "Status": "Aktif" if getattr(r, "is_active", True) else "Nonaktif",
        })

    if not data:
        data = [{"No":"","Jenis":"","Divisi":"","Kategori":"","Nominal (Rp)":"","Status":""}]

    df = pd.DataFrame(data, columns=["No","Jenis","Divisi","Kategori","Nominal (Rp)","Status"])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Insentif")
    bio.seek(0)

    parts = []
    if jenis_q: parts.append(f"jenis-{jenis_q}")
    if divisi_q: parts.append(f"divisi-{divisi_q}")
    if kategori_q: parts.append(f"kategori-{kategori_q}")
    if active_q is not None: parts.append("aktif" if active_q else "nonaktif")
    suffix = "_".join(parts) if parts else "semua"

    return send_file(
        bio,
        as_attachment=True,
        download_name=f"Insentif_{suffix}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ===================== OVERRIDE PER-SALESMAN CRUD =====================
@app.route("/insentif/override")
def override_list():
    rows = db.session.query(InsentifOverride, Salesman)\
            .join(Salesman, InsentifOverride.salesman_id==Salesman.id)\
            .order_by(Salesman.name.asc(), InsentifOverride.jenis.asc(), InsentifOverride.bulan.asc()).all()
    try:
        return render_template("override_list.html", rows=rows)
    except:
        return render_template_string("""
        <div class="container p-4">
          <h3>Override Insentif per Sales</h3>
          <p><a href="{{ url_for('override_new') }}">Tambah Override</a> | <a href="{{ url_for('insentif_list') }}">Kembali ke Rate Umum</a></p>
          <table border="1" cellpadding="6" cellspacing="0">
            <tr><th>Nama Sales</th><th>Jenis</th><th>Bulan</th><th>Nominal</th><th>Aktif</th><th>Aksi</th></tr>
            {% for r,s in rows %}
            <tr>
              <td>{{ s.name }}</td>
              <td>{{ r.jenis }}</td>
              <td>{{ r.bulan or '—' }}</td>
              <td>{{ "{:,}".format(r.amount) }}</td>
              <td>{{ 'Ya' if r.is_active else 'Tidak' }}</td>
              <td>
                <a href="{{ url_for('override_edit', id=r.id) }}">Edit</a> |
                <form action="{{ url_for('override_delete', id=r.id) }}" method="post" style="display:inline" onsubmit="return confirm('Hapus?')">
                  <button type="submit">Delete</button>
                </form>
              </td>
            </tr>
            {% endfor %}
          </table>
        </div>
        """, rows=rows)

@app.route("/insentif/override/new", methods=["GET","POST"])
def override_new():
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method=="POST":
        r = InsentifOverride(
            salesman_id = int(request.form["salesman_id"]),
            jenis = request.form["jenis"],
            bulan = (request.form.get("bulan") or None),
            amount = int(request.form.get("amount") or 0),
            is_active = bool(request.form.get("is_active"))
        )
        db.session.add(r); db.session.commit()
        clear_rate_caches()
        flash("✅ Override dibuat","success")
        return redirect(url_for("override_list"))
    try:
        return render_template("override_form.html", salesmen=salesmen, row=None)
    except:
        return render_template_string(_inline_override_form(), salesmen=salesmen, row=None)

@app.route("/insentif/override/<int:id>/edit", methods=["GET","POST"])
def override_edit(id):
    row = InsentifOverride.query.get_or_404(id)
    salesmen = Salesman.query.order_by(Salesman.name.asc()).all()
    if request.method=="POST":
        row.salesman_id = int(request.form["salesman_id"])
        row.jenis = request.form["jenis"]
        row.bulan = (request.form.get("bulan") or None)
        row.amount = int(request.form.get("amount") or 0)
        row.is_active = bool(request.form.get("is_active"))
        db.session.commit()
        clear_rate_caches()
        flash("✅ Override diperbarui","success")
        return redirect(url_for("override_list"))
    try:
        return render_template("override_form.html", row=row, salesmen=salesmen)
    except:
        return render_template_string(_inline_override_form(), row=row, salesmen=salesmen)

@app.route("/insentif/override/<int:id>/delete", methods=["POST"])
def override_delete(id):
    row = InsentifOverride.query.get_or_404(id)
    db.session.delete(row); db.session.commit()
    clear_rate_caches()
    flash("✅ Override dihapus","success")
    return redirect(url_for("override_list"))

def _inline_override_form():
    return """
    <div class="container p-4">
      <h3>{{ 'Edit' if row else 'Tambah' }} Override Insentif</h3>
      <form method="post">
        <label>Salesman</label>
        <select name="salesman_id" required>
          {% for s in salesmen %}
            <option value="{{ s.id }}" {{ 'selected' if row and row.salesman_id==s.id else '' }}>{{ s.name }} ({{ s.divisi }} / {{ s.kategori }})</option>
          {% endfor %}
        </select><br><br>

        <label>Jenis</label>
        <select name="jenis" required>
          {% set jenis_opts = ['penjualan','tagihan','ca','noo_item','fl','ao','call'] %}
          {% for j in jenis_opts %}
            <option value="{{j}}" {{ 'selected' if row and row.jenis==j else '' }}>{{ j }}</option>
          {% endfor %}
        </select><br><br>

        <label>Bulan (opsional, format YYYY-MM)</label>
        <input type="text" name="bulan" value="{{ row.bulan if row else '' }}" placeholder="2025-10"><br><br>

        <label>Amount (Rp)</label>
        <input type="number" name="amount" value="{{ row.amount if row else 0 }}" required><br><br>

        <label><input type="checkbox" name="is_active" {{ 'checked' if (row and row.is_active) or (not row) else '' }}> Active</label><br><br>

        <button type="submit">Simpan</button>
        <a href="{{ url_for('override_list') }}">Batal</a>
      </form>
    </div>
    """
# ---------- USER CRUD ----------
@app.route("/users")
def user_list():
    users = User.query.order_by(User.id.asc()).all()
    return render_template("user_list.html", users=users)


@app.route("/users/new", methods=["GET", "POST"])
def user_new():
    if request.method == "POST":
        username = request.form["username"].strip()
        full_name = request.form["full_name"].strip()
        role = request.form["role"]
        password = request.form["password"]

        if not username or not password:
            flash("Username dan Password wajib diisi.", "info")
            return redirect(url_for("user_new"))

        if User.query.filter_by(username=username).first():
            flash("Username sudah terdaftar.", "info")
            return redirect(url_for("user_new"))

        u = User(username=username, full_name=full_name, role=role, is_active=True)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash("User berhasil ditambahkan.", "success")
        return redirect(url_for("user_list"))
    return render_template("user_form.html", user=None)


@app.route("/users/edit/<int:id>", methods=["GET", "POST"])
def user_edit(id):
    user = User.query.get_or_404(id)
    if request.method == "POST":
        user.username = request.form["username"].strip()
        user.full_name = request.form["full_name"].strip()
        user.role = request.form["role"]
        password = request.form["password"].strip()

        if password:
            user.set_password(password)

        db.session.commit()
        flash("User berhasil diperbarui.", "success")
        return redirect(url_for("user_list"))
    return render_template("user_form.html", user=user)


@app.route("/users/delete/<int:id>", methods=["POST"])
def user_delete(id):
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash("User berhasil dihapus.", "info")
    return redirect(url_for("user_list"))



# ===================== THRESHOLDS (SETTING PERSENTASE) =====================
@app.route("/thresholds", methods=["GET"])
def thresholds_list():
    rows = Threshold.query.order_by(Threshold.kind.asc(), Threshold.divisi.asc()).all()
    return render_template("thresholds_list.html", rows=rows)


@app.route("/thresholds/edit", methods=["GET", "POST"])
def thresholds_form():
    u = session.get("user")
    if not u or u.get("role") != "superadmin":
        flash("Akses hanya untuk Super Admin.", "danger")
        return redirect(url_for("dashboard"))

    kinds = ["password_awal", "penjualan", "tagihan"]
    divs = ["MP", "AVIAN"]
    pairs = [(k, d) for k in kinds for d in divs]

    if request.method == "POST":
        for kind, div in pairs:
            field_name = f"{kind}_{div}"
            val = (request.form.get(field_name) or "").strip()
            row = Threshold.query.filter_by(kind=kind, divisi=div).first()
            if not row:
                row = Threshold(kind=kind, divisi=div)
                db.session.add(row)
            try:
                row.percent = float(val) if val != "" else None
            except:
                row.percent = None
        db.session.commit()
        flash("✅ Persentase threshold berhasil disimpan.", "success")
        return redirect(url_for("thresholds_list"))

    current = {}
    for k, d in pairs:
        r = Threshold.query.filter_by(kind=k, divisi=d).first()
        current[f"{k}_{d}"] = r.percent if r else None

    return render_template("thresholds_form.html", current=current)



# ===================== CLI =====================
@app.cli.command("init-db")
def init_db():
    """Inisialisasi database + seed data awal"""
    db.create_all()
    print("📦 Database initialized.")

    # Pastikan superadmin selalu ada
    ensure_super_admin()

    # Seed salesman contoh (opsional)
    if Salesman.query.count() == 0:
        db.session.add(Salesman(name="Contoh MP", area="Sukabumi", divisi="MP", kategori="Bronze"))
        db.session.add(Salesman(name="Contoh Avian", area="Depok", divisi="AVIAN", kategori="Silver"))
        db.session.commit()
        print("👥 Seeded sample salesmen.")

    # Seed insentif rate (sekali saja)
    if InsentifRate.query.count() == 0:
        seeds = [
            # Penjualan MP
            ("penjualan","MP","Bronze", 500_000),
            ("penjualan","MP","Silver", 700_000),
            ("penjualan","MP","Gold",   900_000),
            ("penjualan","MP","Platinum",1_100_000),
            # Penjualan AVIAN
            ("penjualan","AVIAN","Bronze", 400_000),
            ("penjualan","AVIAN","Silver", 600_000),
            ("penjualan","AVIAN","Gold",   800_000),
            ("penjualan","AVIAN","Platinum",1_000_000),
            # Tagihan
            ("tagihan","MP","Bronze", 500_000),
            ("tagihan","MP","Silver", 700_000),
            ("tagihan","MP","Gold",   900_000),
            ("tagihan","MP","Platinum",1_100_000),
            ("tagihan","AVIAN","Bronze", 400_000),
            ("tagihan","AVIAN","Silver", 600_000),
            ("tagihan","AVIAN","Gold",   800_000),
            ("tagihan","AVIAN","Platinum",1_000_000),
            # Komponen lain
            ("ca", None, None, 100_000),
            ("noo_item", None, None, 50_000),
            ("fl", None, None, 100_000),
            ("ao", None, None, 100_000),
            # Optional: aktifkan kalau mau
            # ("call", None, None, 0),
        ]
        for jenis, div, kat, amt in seeds:
            db.session.add(InsentifRate(jenis=jenis, divisi=div, kategori=kat, amount=amt))
        db.session.commit()
        print("💰 Seeded insentif rates.")

    print("✅ Database setup complete.")

        

# ===================== MAIN =====================
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        ensure_super_admin()
    app.run(debug=False)
